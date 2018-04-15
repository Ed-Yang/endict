# -*- coding: utf-8 -*-

import csv
from urllib.request import urlopen
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl

href = "http://www.taiwantestcentral.com/WordList/{}"
# cells[4].contents[0].attrs['href']
html = urlopen("http://www.taiwantestcentral.com/WordList/BCTWordList.aspx?CategoryID=12") 

def collect_words(rows):

    wlist = []

    for row in rows[1:]:
        cells = row.findAll(['td', 'th'])
        txt0 = cells[0].get_text()
        txt2 = cells[2].get_text()
        txt4 = cells[4].get_text()
        try:
            lnk = cells[4].contents[0].attrs['href']
        except:
            lnk = ''

        wlist.append([txt0, txt2, txt4, lnk])

    return wlist

def write_csv(filename, wlist, words=9, skip_lvl=50):
    csvFile = open(filename, 'wt')
    writer = csv.writer(csvFile, delimiter=',')

    wc = 0
    csvRow = []
    skip = 0

    try:
        for row in wlist:
            try:
                if row[2] and int(row[2]) > skip_lvl:
                    print("skip --> ", row[0], row[2])
                    skip = skip + 1
                    continue
            except:
                print('here')

            csvRow.append(row[0])
            wc = wc + 1

            if wc == words:
                writer.writerow(csvRow) 
                wc = 0
                csvRow = []

    finally:
        if wc != 0:
            writer.writerow(csvRow) 
        csvFile.close()
        print('skip words = ', skip)


def write_xlsx(filename, wlist):

    book = openpyxl.Workbook()
    sheet = book.create_sheet('1200-words')
    max_length = 0

    for ro, line in enumerate(wlist):
        
        r = ro+1
        sheet.cell(row=r, column=1).value = line[0]
        m = sheet.cell(row=r, column=2).value = line[1]

        if len(m) > max_length:
            max_length = len(m)

        sheet.cell(row=r, column=3).value = line[2]
        sheet.cell(row=r, column=3).hyperlink = href.format(line[3])
        
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions['B'].width = adjusted_width
    book.save('en-1200.xlsx')

if __name__ == '__main__':

    bsObj = BeautifulSoup(html.read(), "lxml")
    print(bsObj.h1)

    table = bsObj.findAll("table",{"class":"WordList"})[0]
    HtmlRows = table.findAll("tr")
    WordList = collect_words(HtmlRows)
    write_csv("en-1200.csv", WordList)
    write_xlsx("en-1200.csv", WordList)
    print('Done')

